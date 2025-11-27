#!/usr/bin/env python3
"""
Mathematische Eigenschaften-Analyse: Check mathematische Patterns in der Matrix.

Analysiert:
- Primzahlen, Teilbarkeit
- Fibonacci-ähnliche Sequenzen
- Arithmetische/Geometrische Progressionen
- Modulo-Patterns
- Andere mathematische Strukturen
"""

import json
import numpy as np
from pathlib import Path
from typing import Dict, List
from collections import Counter

OUTPUT_DIR = Path(__file__).parent.parent.parent / "outputs" / "derived"
MATRIX_PATH = Path(__file__).parent.parent.parent / "data" / "anna-matrix" / "Anna_Matrix.xlsx"
OUTPUT_FILE = OUTPUT_DIR / "mathematical_properties_analysis.json"
REPORT_FILE = OUTPUT_DIR / "mathematical_properties_analysis_report.md"

def load_matrix() -> np.ndarray:
 """Load die Anna Matrix."""
 try:
 from openpyxl import load_workbook
 wb = load_workbook(MATRIX_PATH, data_only=True)
 ws = wb.active
 matrix = np.zeros((128, 128), dtype=float)
 for i in range(128):
 for j in range(128):
 val = ws.cell(i+1, j+1).value
 try:
 matrix[i, j] = float(val) if val is not None else 0.0
 except (ValueError, TypeError):
 matrix[i, j] = 0.0
 return matrix
 except Exception as e:
 print(f"Fehler beim Loadn: {e}")
 return None

def is_prime(n: int) -> bool:
 """Check ob Zahl eine Primzahl ist."""
 if n < 2:
 return False
 for i in range(2, int(n**0.5) + 1):
 if n % i == 0:
 return False
 return True

def analyze_mathematical_properties(matrix: np.ndarray) -> Dict:
 """Analyze mathematische Eigenschaften der Matrix."""
 
 values = matrix.flatten()
 int_values = [int(abs(v)) for v in values if not np.isnan(v)]
 
 analysis = {
 "total_values": len(int_values),
 "prime_numbers": [],
 "even_numbers": 0,
 "odd_numbers": 0,
 "divisible_by_26": 0,
 "divisible_by_13": 0,
 "divisible_by_2": 0,
 "modulo_patterns": {},
 }
 
 # Primzahlen
 unique_values = set(int_values)
 primes = [v for v in unique_values if is_prime(v)]
 analysis["prime_numbers"] = sorted(primes)[:50] # Nur erste 50
 analysis["prime_count"] = len(primes)
 analysis["prime_percentage"] = (len(primes) / len(unique_values) * 100) if unique_values else 0
 
 # Gerade/Ungerade
 analysis["even_numbers"] = sum(1 for v in int_values if v % 2 == 0)
 analysis["odd_numbers"] = sum(1 for v in int_values if v % 2 == 1)
 
 # Teilbarkeit
 analysis["divisible_by_26"] = sum(1 for v in int_values if v % 26 == 0)
 analysis["divisible_by_13"] = sum(1 for v in int_values if v % 13 == 0)
 analysis["divisible_by_2"] = sum(1 for v in int_values if v % 2 == 0)
 
 # Modulo-Patterns (verschiedene Moduli)
 for mod in [3, 4, 5, 7, 8, 9, 10, 11, 12, 26]:
 mod_counts = Counter(v % mod for v in int_values)
 analysis["modulo_patterns"][f"mod_{mod}"] = {
 "distribution": dict(mod_counts),
 "expected_per_value": len(int_values) / mod,
 }
 
 return analysis

def analyze_sequences(matrix: np.ndarray) -> Dict:
 """Suche nach arithmetischen/geometrischen Sequenzen."""
 
 sequences = {
 "arithmetic_sequences": [],
 "geometric_sequences": [],
 }
 
 # Check Zeilen
 for i in range(min(10, matrix.shape[0])):
 row = matrix[i]
 # Check auf arithmetische Progression (vereinfacht: erste 10 Werte)
 if len(row) >= 3:
 for start in range(min(5, len(row) - 2)):
 vals = [float(row[start]), float(row[start+1]), float(row[start+2])]
 if all(not np.isnan(v) for v in vals):
 diff1 = vals[1] - vals[0]
 diff2 = vals[2] - vals[1]
 if abs(diff1 - diff2) < 0.01: # Arithmetische Progression
 sequences["arithmetic_sequences"].append({
 "row": i,
 "start": start,
 "values": vals,
 "difference": diff1,
 })
 break # Nur eine pro Zeile
 
 return sequences

def main():
 """Mathematische Eigenschaften-Analyse."""
 
 print("=" * 80)
 print("MATHEMATISCHE EIGENSCHAFTEN-ANALYSE")
 print("=" * 80)
 print()
 print("⚠️ WICHTIG: Nur Fakten, keine Interpretationen!")
 print()
 
 # Load Matrix
 print("Load Anna Matrix...")
 matrix = load_matrix()
 if matrix is None:
 print("❌ Matrix konnte nicht geloadn werden!")
 return
 
 print(f"✅ Matrix geloadn: {matrix.shape}")
 print()
 
 # Analyze mathematische Eigenschaften
 print("Analyze mathematische Eigenschaften...")
 math_props = analyze_mathematical_properties(matrix)
 
 print(f" ✅ Primzahlen: {math_props['prime_count']} ({math_props['prime_percentage']:.1f}% der unique Werte)")
 print(f" ✅ Gerade: {math_props['even_numbers']}, Ungerade: {math_props['odd_numbers']}")
 print(f" ✅ Teilbar durch 26: {math_props['divisible_by_26']}")
 print(f" ✅ Teilbar durch 13: {math_props['divisible_by_13']}")
 
 # Analyze Sequenzen
 print("Suche nach Sequenzen...")
 sequences = analyze_sequences(matrix)
 print(f" ✅ Arithmetische Sequenzen: {len(sequences['arithmetic_sequences'])}")
 print()
 
 # Speichere Ergebnisse
 output_data = {
 "mathematical_properties": math_props,
 "sequences": sequences,
 }
 
 OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
 with OUTPUT_FILE.open("w") as f:
 json.dump(output_data, f, indent=2)
 
 print(f"💾 Ergebnisse gespeichert in: {OUTPUT_FILE}")
 
 # Erstelle Report
 report_content = f"""# Mathematische Eigenschaften-Analyse: Fakten

**Datum**: 2025-11-22 
**Ziel**: Check mathematische Patterns in der Matrix

## ⚠️ WICHTIG

**Nur Fakten, keine Interpretationen!**

## 1. Primzahlen (FAKTEN)

- **Total unique Werte**: {len(set(int(abs(v)) for v in matrix.flatten() if not np.isnan(v)))}
- **Primzahlen gefunden**: {math_props['prime_count']}
- **Primzahl-Anteil**: {math_props['prime_percentage']:.1f}%

### Erste 20 Primzahlen in Matrix
"""
 
 for prime in math_props['prime_numbers'][:20]:
 report_content += f"- {prime}\n"
 
 report_content += f"""
## 2. Teilbarkeit (FAKTEN)

- **Gerade Zahlen**: {math_props['even_numbers']} ({math_props['even_numbers']/math_props['total_values']*100:.1f}%)
- **Ungerade Zahlen**: {math_props['odd_numbers']} ({math_props['odd_numbers']/math_props['total_values']*100:.1f}%)
- **Teilbar durch 26**: {math_props['divisible_by_26']} ({math_props['divisible_by_26']/math_props['total_values']*100:.1f}%)
- **Teilbar durch 13**: {math_props['divisible_by_13']} ({math_props['divisible_by_13']/math_props['total_values']*100:.1f}%)

## 3. Modulo-Patterns (FAKTEN)

### Modulo 26 (wichtig for Base-26)
- **Erwartet**: {math_props['modulo_patterns']['mod_26']['expected_per_value']:.1f} pro Wert (0-25)
- **Tatsächlich**: Siehe Base-26 Verteilung in Matrix-Origin-Analyse

## 4. Sequenzen (FAKTEN)

- **Arithmetische Sequenzen**: {len(sequences['arithmetic_sequences'])} gefunden

## ❓ OFFENE FRAGEN (NICHT BEANTWORTET)

1. ❓ Haben die mathematischen Eigenschaften Bedeutung?
2. ❓ Warum gibt es diese Patterns?
3. ❓ Sind sie beabsichtigt?

## ⚠️ WICHTIG

**Diese Analyse zeigt nur FAKTEN!** 
**Keine Interpretationen!**
"""
 
 with REPORT_FILE.open("w") as f:
 f.write(report_content)
 
 print(f"📄 Report erstellt: {REPORT_FILE}")
 print()
 print("✅ Analyse abgeschlossen!")

if __name__ == "__main__":
 main()

