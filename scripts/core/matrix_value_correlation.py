#!/usr/bin/env python3
"""
Matrix-Wert-Korrelation: Check ob Matrix-Werte mit Identity-Eigenschaften korrelieren.

Analysiert:
- Korrelation zwischen Matrix-Werten und Seed-Eigenschaften
- Korrelation zwischen Koordinaten und Identity-Eigenschaften
- Mathematische Zusammenhänge
"""

import json
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple
from collections import defaultdict, Counter

OUTPUT_DIR = Path(__file__).parent.parent.parent / "outputs" / "derived"
COMPREHENSIVE_SCAN = OUTPUT_DIR / "comprehensive_matrix_scan.json"
MATRIX_PATH = Path(__file__).parent.parent.parent / "data" / "anna-matrix" / "Anna_Matrix.xlsx"
OUTPUT_FILE = OUTPUT_DIR / "matrix_value_correlation.json"
REPORT_FILE = OUTPUT_DIR / "matrix_value_correlation_report.md"

def load_matrix() -> np.ndarray:
 """Load die Anna Matrix."""
 try:
 from openpyxl import load_workbook
 wb = load_workbook(MATRIX_PATH, data_only=True)
 ws = wb.active
 matrix = np.zeros((128, 128), dtype=float)
 for i in range(128):
 for j in range(128):
 val = ws.cell(i+1, j+1).value
 try:
 matrix[i, j] = float(val) if val is not None else 0.0
 except (ValueError, TypeError):
 matrix[i, j] = 0.0
 return matrix
 except Exception as e:
 print(f"Fehler beim Loadn: {e}")
 return None

def analyze_matrix_values_at_coordinates(matrix: np.ndarray, comprehensive_data: Dict) -> Dict:
 """Analyze Matrix-Werte an den Koordinaten der Identities."""
 
 results = comprehensive_data.get("results", [])
 
 coordinate_values = []
 coordinate_base26 = []
 
 for pattern_result in results:
 pattern_name = pattern_result.get("pattern_name", "unknown")
 identities = pattern_result.get("identities", [])
 
 for identity_record in identities:
 path = identity_record.get("path", [])
 identity = identity_record.get("identity", "")
 
 if path and identity:
 # Extrahiere Matrix-Werte an diesen Koordinaten
 values = []
 base26_values = []
 
 for coord in path:
 if len(coord) == 2:
 row, col = coord[0], coord[1]
 if 0 <= row < 128 and 0 <= col < 128:
 val = matrix[row, col]
 values.append(float(val))
 base26_values.append(int(val) % 26)
 
 if values:
 coordinate_values.append({
 "identity": identity[:40] + "...",
 "pattern": pattern_name,
 "values": values,
 "mean": float(np.mean(values)),
 "std": float(np.std(values)),
 "min": float(np.min(values)),
 "max": float(np.max(values)),
 "base26_values": base26_values,
 })
 
 # Analyze Patterns
 all_values = []
 all_base26 = []
 pattern_stats = defaultdict(list)
 
 for cv in coordinate_values:
 all_values.extend(cv["values"])
 all_base26.extend(cv["base26_values"])
 pattern_stats[cv["pattern"]].append(cv["mean"])
 
 analysis = {
 "total_coordinate_sets": len(coordinate_values),
 "total_values_analyzed": len(all_values),
 "overall_stats": {
 "mean": float(np.mean(all_values)) if all_values else 0,
 "std": float(np.std(all_values)) if all_values else 0,
 "min": float(np.min(all_values)) if all_values else 0,
 "max": float(np.max(all_values)) if all_values else 0,
 },
 "base26_distribution": dict(Counter(all_base26)),
 "pattern_means": {k: float(np.mean(v)) for k, v in pattern_stats.items()},
 }
 
 return analysis

def main():
 """Matrix-Wert-Korrelations-Analyse."""
 
 print("=" * 80)
 print("MATRIX-WERT-KORRELATIONS-ANALYSE")
 print("=" * 80)
 print()
 print("⚠️ WICHTIG: Nur Fakten, keine Interpretationen!")
 print()
 
 # Load Matrix
 print("Load Anna Matrix...")
 matrix = load_matrix()
 if matrix is None:
 print("❌ Matrix konnte nicht geloadn werden!")
 return
 
 print(f"✅ Matrix geloadn: {matrix.shape}")
 
 # Load Comprehensive Scan
 if not COMPREHENSIVE_SCAN.exists():
 print(f"❌ Datei nicht gefunden: {COMPREHENSIVE_SCAN}")
 return
 
 print("Load Comprehensive Scan Daten...")
 with COMPREHENSIVE_SCAN.open() as f:
 comprehensive_data = json.load(f)
 
 print(f"✅ {len(comprehensive_data.get('results', []))} Patterns geloadn")
 print()
 
 # Analyze Matrix-Werte an Koordinaten
 print("Analyze Matrix-Werte an Identity-Koordinaten...")
 correlation = analyze_matrix_values_at_coordinates(matrix, comprehensive_data)
 
 print(f" ✅ {correlation['total_coordinate_sets']} Koordinaten-Sets analysiert")
 print(f" ✅ {correlation['total_values_analyzed']:,} Matrix-Werte analysiert")
 print(f" ✅ Mean: {correlation['overall_stats']['mean']:.2f}, Std: {correlation['overall_stats']['std']:.2f}")
 print()
 
 # Speichere Ergebnisse
 OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
 with OUTPUT_FILE.open("w") as f:
 json.dump(correlation, f, indent=2)
 
 print(f"💾 Ergebnisse gespeichert in: {OUTPUT_FILE}")
 
 # Erstelle Report
 report_content = f"""# Matrix-Wert-Korrelations-Analyse: Fakten

**Datum**: 2025-11-22 
**Ziel**: Check Korrelationen zwischen Matrix-Werten und Identities

## ⚠️ WICHTIG

**Nur Fakten, keine Interpretationen!**

## 1. Matrix-Werte an Identity-Koordinaten (FAKTEN)

- **Total Koordinaten-Sets**: {correlation['total_coordinate_sets']}
- **Total Matrix-Werte analysiert**: {correlation['total_values_analyzed']:,}

### Gesamt-Statistiken
- **Mean**: {correlation['overall_stats']['mean']:.2f}
- **Std**: {correlation['overall_stats']['std']:.2f}
- **Min**: {correlation['overall_stats']['min']:.1f}
- **Max**: {correlation['overall_stats']['max']:.1f}

### Pattern-spezifische Means
"""
 
 for pattern, mean_val in correlation.get("pattern_means", {}).items():
 report_content += f"- **{pattern}**: {mean_val:.2f}\n"
 
 report_content += """
## ❓ OFFENE FRAGEN (NICHT BEANTWORTET)

1. ❓ Gibt es Korrelationen zwischen Matrix-Werten und Identity-Eigenschaften?
2. ❓ Warum haben manche Patterns unterschiedliche Mean-Werte?
3. ❓ Was bedeuten die Matrix-Werte an den Koordinaten?

## ⚠️ WICHTIG

**Diese Analyse zeigt nur FAKTEN!** 
**Keine Interpretationen!**
"""
 
 with REPORT_FILE.open("w") as f:
 f.write(report_content)
 
 print(f"📄 Report erstellt: {REPORT_FILE}")
 print()
 print("✅ Analyse abgeschlossen!")

if __name__ == "__main__":
 main()

