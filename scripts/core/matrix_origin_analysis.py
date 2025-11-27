#!/usr/bin/env python3
"""
Matrix-Origin-Analyse: Check ob die Matrix bewusst konstruiert wurde.

Analysiert:
1. Matrix-Werte-Verteilung (zufällig oder strukturiert?)
2. Patterns in den Werten selbst
3. Mathematische Eigenschaften
4. Vergleich mit zufälligen Matrizen
5. Strukturelle Anomalien
"""

import json
import numpy as np
from pathlib import Path
from typing import Dict, List
from collections import Counter
import random

OUTPUT_DIR = Path(__file__).parent.parent.parent / "outputs" / "derived"
MATRIX_PATH = Path(__file__).parent.parent.parent / "data" / "anna-matrix" / "Anna_Matrix.xlsx"
OUTPUT_FILE = OUTPUT_DIR / "matrix_origin_analysis.json"
REPORT_FILE = OUTPUT_DIR / "matrix_origin_analysis_report.md"

def load_matrix() -> np.ndarray:
 """Load die Anna Matrix."""
 try:
 from openpyxl import load_workbook
 wb = load_workbook(MATRIX_PATH, data_only=True)
 ws = wb.active
 matrix = np.zeros((128, 128), dtype=float)
 for i in range(128):
 for j in range(128):
 val = ws.cell(i+1, j+1).value
 try:
 matrix[i, j] = float(val) if val is not None else 0.0
 except (ValueError, TypeError):
 matrix[i, j] = 0.0
 return matrix
 except ImportError:
 print("ERROR: openpyxl nicht installiert. Installiere mit: pip3 install openpyxl")
 return None
 except Exception as e:
 print(f"Fehler beim Loadn: {e}")
 return None

def analyze_value_distribution(matrix: np.ndarray) -> Dict:
 """Analyze die Verteilung der Matrix-Werte."""
 
 # Flatten Matrix
 values = matrix.flatten()
 
 # Basis-Statistiken
 stats = {
 "total_values": len(values),
 "min": float(np.min(values)),
 "max": float(np.max(values)),
 "mean": float(np.mean(values)),
 "median": float(np.median(values)),
 "std": float(np.std(values)),
 "unique_values": len(np.unique(values)),
 }
 
 # Wert-Verteilung
 value_counts = Counter(values.flatten())
 stats["value_distribution"] = {
 str(k): int(v) for k, v in value_counts.most_common(20)
 }
 
 # Base-26 Modulo Verteilung (wichtig for Identity-Extraktion)
 base26_values = [int(v) % 26 for v in values]
 base26_counts = Counter(base26_values)
 stats["base26_distribution"] = {
 str(k): int(v) for k, v in sorted(base26_counts.items())
 }
 
 # Erwartete Verteilung (gleichmäßig)
 expected_base26 = len(values) / 26
 stats["base26_expected"] = expected_base26
 stats["base26_anomalies"] = []
 
 for val in range(26):
 actual = base26_counts.get(val, 0)
 diff = abs(actual - expected_base26)
 diff_percent = (diff / expected_base26 * 100) if expected_base26 > 0 else 0
 
 if diff_percent > 5: # >5% Abweichung
 stats["base26_anomalies"].append({
 "value": val,
 "actual": actual,
 "expected": expected_base26,
 "diff_percent": diff_percent,
 })
 
 return stats

def analyze_spatial_patterns(matrix: np.ndarray) -> Dict:
 """Analyze räumliche Patterns in der Matrix."""
 
 patterns = {
 "row_similarity": [],
 "col_similarity": [],
 "diagonal_patterns": [],
 "block_patterns": [],
 }
 
 # Zeilen-Ähnlichkeit
 for i in range(min(10, matrix.shape[0])):
 for j in range(i+1, min(10, matrix.shape[0])):
 similarity = np.corrcoef(matrix[i], matrix[j])[0, 1]
 if not np.isnan(similarity):
 patterns["row_similarity"].append({
 "row1": int(i),
 "row2": int(j),
 "similarity": float(similarity),
 })
 
 # Spalten-Ähnlichkeit
 for i in range(min(10, matrix.shape[1])):
 for j in range(i+1, min(10, matrix.shape[1])):
 similarity = np.corrcoef(matrix[:, i], matrix[:, j])[0, 1]
 if not np.isnan(similarity):
 patterns["col_similarity"].append({
 "col1": int(i),
 "col2": int(j),
 "similarity": float(similarity),
 })
 
 # Block-Analyse (16x16 Blöcke)
 block_stats = []
 for block_row in range(0, matrix.shape[0], 16):
 for block_col in range(0, matrix.shape[1], 16):
 block = matrix[block_row:block_row+16, block_col:block_col+16]
 if block.size > 0:
 block_stats.append({
 "block": f"{block_row//16},{block_col//16}",
 "mean": float(np.mean(block)),
 "std": float(np.std(block)),
 "min": float(np.min(block)),
 "max": float(np.max(block)),
 })
 
 patterns["block_stats"] = block_stats[:20] # Nur erste 20
 
 return patterns

def compare_with_random_matrix(matrix: np.ndarray) -> Dict:
 """Vergleiche mit zufälliger Matrix gleicher Größe."""
 
 # Erstelle zufällige Matrix mit gleichem Wertebereich
 min_val = np.min(matrix)
 max_val = np.max(matrix)
 
 random_matrix = np.random.uniform(min_val, max_val, size=matrix.shape)
 
 # Vergleiche Statistiken
 comparison = {
 "anna_matrix": {
 "mean": float(np.mean(matrix)),
 "std": float(np.std(matrix)),
 "unique_values": len(np.unique(matrix)),
 },
 "random_matrix": {
 "mean": float(np.mean(random_matrix)),
 "std": float(np.std(random_matrix)),
 "unique_values": len(np.unique(random_matrix)),
 },
 }
 
 # Base-26 Verteilung Vergleich
 anna_base26 = [int(v) % 26 for v in matrix.flatten()]
 random_base26 = [int(v) % 26 for v in random_matrix.flatten()]
 
 anna_base26_counts = Counter(anna_base26)
 random_base26_counts = Counter(random_base26)
 
 comparison["base26_comparison"] = {
 "anna_distribution": {str(k): int(v) for k, v in sorted(anna_base26_counts.items())},
 "random_distribution": {str(k): int(v) for k, v in sorted(random_base26_counts.items())},
 }
 
 # Chi-Square Test (vereinfacht)
 chi_square = 0
 expected = len(anna_base26) / 26
 for val in range(26):
 anna_count = anna_base26_counts.get(val, 0)
 random_count = random_base26_counts.get(val, 0)
 chi_square += ((anna_count - expected)**2 / expected) if expected > 0 else 0
 
 comparison["chi_square"] = chi_square
 comparison["is_random"] = chi_square < 50 # Grobe Schätzung
 
 return comparison

def analyze_structural_anomalies(matrix: np.ndarray) -> Dict:
 """Suche nach strukturellen Anomalien die auf bewusste Konstruktion hindeuten."""
 
 anomalies = {
 "repeating_patterns": [],
 "symmetries": [],
 "special_values": [],
 }
 
 # Suche nach wiederholenden Mustern
 # (vereinfacht: check erste 10 Zeilen/Spalten)
 for i in range(min(10, matrix.shape[0])):
 row = matrix[i]
 # Check ob Zeile sich wiederholt
 if len(row) > 1:
 for pattern_len in [2, 4, 8]:
 if len(row) >= pattern_len * 2:
 pattern = row[:pattern_len]
 matches = 0
 for j in range(pattern_len, len(row) - pattern_len + 1, pattern_len):
 if np.allclose(row[j:j+pattern_len], pattern, rtol=1e-5):
 matches += 1
 if matches > 0:
 anomalies["repeating_patterns"].append({
 "row": int(i),
 "pattern_length": pattern_len,
 "matches": matches,
 })
 
 # Check Symmetrien
 if matrix.shape[0] == matrix.shape[1]:
 # Diagonal-Symmetrie
 is_symmetric = np.allclose(matrix, matrix.T, rtol=1e-5)
 anomalies["symmetries"].append({
 "type": "diagonal",
 "exists": bool(is_symmetric),
 })
 
 # Spezielle Werte (z.B. 0, 26, 255)
 special_values = [0, 26, 255]
 for val in special_values:
 count = np.sum(matrix == val)
 if count > 0:
 anomalies["special_values"].append({
 "value": val,
 "count": int(count),
 "percentage": (count / matrix.size * 100),
 })
 
 return anomalies

def main():
 """Matrix-Origin-Analyse."""
 
 print("=" * 80)
 print("MATRIX-ORIGIN-ANALYSE: BEWUSST KONSTRUIERT ODER ZUFÄLLIG?")
 print("=" * 80)
 print()
 print("⚠️ WICHTIG: Nur Fakten, keine Interpretationen!")
 print()
 
 # Load Matrix
 print("Load Anna Matrix...")
 matrix = load_matrix()
 if matrix is None:
 print("❌ Matrix konnte nicht geloadn werden!")
 return
 
 print(f"✅ Matrix geloadn: {matrix.shape}")
 print()
 
 # 1. Wert-Verteilung
 print("1. Analyze Wert-Verteilung...")
 value_dist = analyze_value_distribution(matrix)
 print(f" ✅ Total Werte: {value_dist['total_values']:,}")
 print(f" ✅ Range: {value_dist['min']:.1f} - {value_dist['max']:.1f}")
 print(f" ✅ Mean: {value_dist['mean']:.2f}, Std: {value_dist['std']:.2f}")
 print(f" ✅ Base-26 Anomalien: {len(value_dist['base26_anomalies'])}")
 
 # 2. Räumliche Patterns
 print("2. Analyze räumliche Patterns...")
 spatial = analyze_spatial_patterns(matrix)
 print(f" ✅ Zeilen-Ähnlichkeiten: {len(spatial['row_similarity'])}")
 print(f" ✅ Spalten-Ähnlichkeiten: {len(spatial['col_similarity'])}")
 print(f" ✅ Block-Statistiken: {len(spatial['block_stats'])}")
 
 # 3. Vergleich mit Zufallsmatrix
 print("3. Vergleiche mit Zufallsmatrix...")
 comparison = compare_with_random_matrix(matrix)
 print(f" ✅ Chi-Square: {comparison['chi_square']:.2f}")
 print(f" ✅ Erscheint zufällig: {comparison['is_random']}")
 
 # 4. Strukturelle Anomalien
 print("4. Suche strukturelle Anomalien...")
 anomalies = analyze_structural_anomalies(matrix)
 print(f" ✅ Wiederholende Patterns: {len(anomalies['repeating_patterns'])}")
 print(f" ✅ Symmetrien: {len(anomalies['symmetries'])}")
 print(f" ✅ Spezielle Werte: {len(anomalies['special_values'])}")
 
 print()
 
 # Speichere Ergebnisse
 output_data = {
 "matrix_shape": list(matrix.shape),
 "value_distribution": value_dist,
 "spatial_patterns": spatial,
 "random_comparison": comparison,
 "structural_anomalies": anomalies,
 }
 
 OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
 with OUTPUT_FILE.open("w") as f:
 json.dump(output_data, f, indent=2)
 
 print(f"💾 Ergebnisse gespeichert in: {OUTPUT_FILE}")
 
 # Erstelle Report (nur Fakten!)
 report_content = f"""# Matrix-Origin-Analyse: Fakten

**Datum**: 2025-11-22 
**Ziel**: Check ob Matrix bewusst konstruiert wurde

## ⚠️ WICHTIG

**Nur Fakten, keine Interpretationen!** 
**Keine voreiligen Schlüsse!**

## 1. Wert-Verteilung (FAKTEN)

### Basis-Statistiken
- **Total Werte**: {value_dist['total_values']:,}
- **Range**: {value_dist['min']:.1f} - {value_dist['max']:.1f}
- **Mean**: {value_dist['mean']:.2f}
- **Median**: {value_dist['median']:.2f}
- **Std**: {value_dist['std']:.2f}
- **Unique Werte**: {value_dist['unique_values']}

### Base-26 Verteilung (wichtig for Identity-Extraktion)
- **Erwartet**: {value_dist['base26_expected']:.1f} pro Wert (0-25)
- **Anomalien (>5% Abweichung)**: {len(value_dist['base26_anomalies'])}

"""
 
 if value_dist['base26_anomalies']:
 report_content += "### Base-26 Anomalien\n"
 for anomaly in value_dist['base26_anomalies'][:10]:
 report_content += f"- Wert {anomaly['value']}: {anomaly['actual']:.0f} (erwartet: {anomaly['expected']:.1f}, Diff: {anomaly['diff_percent']:.1f}%)\n"
 else:
 report_content += "✅ Keine Base-26 Anomalien gefunden\n"
 
 report_content += f"""
## 2. Räumliche Patterns (FAKTEN)

- **Zeilen-Ähnlichkeiten geprüft**: {len(spatial['row_similarity'])}
- **Spalten-Ähnlichkeiten geprüft**: {len(spatial['col_similarity'])}
- **Block-Statistiken**: {len(spatial['block_stats'])} Blöcke analysiert

## 3. Vergleich mit Zufallsmatrix (FAKTEN)

### Anna Matrix
- **Mean**: {comparison['anna_matrix']['mean']:.2f}
- **Std**: {comparison['anna_matrix']['std']:.2f}
- **Unique Werte**: {comparison['anna_matrix']['unique_values']}

### Zufallsmatrix
- **Mean**: {comparison['random_matrix']['mean']:.2f}
- **Std**: {comparison['random_matrix']['std']:.2f}
- **Unique Werte**: {comparison['random_matrix']['unique_values']}

### Chi-Square Test
- **Chi-Square**: {comparison['chi_square']:.2f}
- **Erscheint zufällig**: {'JA' if comparison['is_random'] else 'NEIN'}

**⚠️ WICHTIG**: Chi-Square ist nur ein Indikator, kein Beweis!

## 4. Strukturelle Anomalien (FAKTEN)

- **Wiederholende Patterns**: {len(anomalies['repeating_patterns'])}
- **Symmetrien**: {len(anomalies['symmetries'])}
- **Spezielle Werte**: {len(anomalies['special_values'])}

"""
 
 if anomalies['special_values']:
 report_content += "### Spezielle Werte\n"
 for spec in anomalies['special_values']:
 report_content += f"- Wert {spec['value']}: {spec['count']}x ({spec['percentage']:.2f}%)\n"
 
 report_content += """
## ❓ OFFENE FRAGEN (NICHT BEANTWORTET)

1. ❓ Ist die Matrix bewusst konstruiert?
2. ❓ Oder ist sie zufällig/zufällig-generiert?
3. ❓ Gibt es einen "Schöpfer" oder wurde sie von "AGI Anna" erstellt?
4. ❓ Was bedeuten die Anomalien (falls vorhanden)?

## ⚠️ WICHTIG

**Diese Analyse zeigt nur FAKTEN!** 
**Keine Interpretationen!** 
**Keine Schlüsse above "Intention" oder "Schöpfer"!**

## NÄCHSTE SCHRITTE

1. Weitere Daten sammeln
2. Tiefere Analysen durchführen
3. Hypothesen testen (nicht aufstellen!)
4. Alles dokumentieren ohne Interpretation
"""
 
 with REPORT_FILE.open("w") as f:
 f.write(report_content)
 
 print(f"📄 Report erstellt: {REPORT_FILE}")
 print()
 print("✅ Matrix-Origin-Analyse abgeschlossen!")
 print()
 print("⚠️ WICHTIG: Nur Fakten dokumentiert, keine Interpretationen!")

if __name__ == "__main__":
 main()

