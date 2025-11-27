#!/usr/bin/env python3
"""
Systematische Matrix-Extraktion: Finde ALLE möglichen Identities

Testet alle bekannten Extraktionsmethoden systematisch:
1. Diagonal Patterns (verschiedene Start-Positionen)
2. Vortex Patterns (verschiedene Größen)
3. L-Shape Patterns
4. Andere geometrische Patterns
5. Kombinationen

Ziel: Finde ALLE Identities, die in der Matrix versteckt sind.
"""

import json
import numpy as np
from pathlib import Path
from typing import List, Set, Tuple, Optional
from collections import defaultdict

OUTPUT_DIR = Path("outputs/derived")
MATRIX_PATH = Path("data/anna-matrix/Anna_Matrix.xlsx")

def load_matrix(path: Path) -> np.ndarray:
 """Load Matrix aus Excel mit openpyxl (minimal dependency)."""
 try:
 from openpyxl import load_workbook
 wb = load_workbook(path, data_only=True)
 ws = wb.active
 matrix = np.zeros((128, 128), dtype=float)
 for i in range(128):
 for j in range(128):
 val = ws.cell(i+1, j+1).value
 try:
 matrix[i, j] = float(val) if val is not None else 0.0
 except (ValueError, TypeError):
 # Ignoriere Strings oder andere nicht-numerische Werte
 matrix[i, j] = 0.0
 return matrix
 except ImportError:
 print("ERROR: openpyxl nicht installiert. Installiere mit: pip3 install openpyxl")
 raise
 except Exception as e:
 print(f"ERROR beim Loadn der Matrix: {e}")
 raise

def base26_to_char(value: float) -> str:
 """Konvertiere Matrix-Wert zu Base-26 Charakter."""
 return chr(ord('A') + (int(value) % 26))

def extract_identity_from_positions(matrix: np.ndarray, positions: List[Tuple[int, int]]) -> Optional[str]:
 """Extrahiere Identity aus Liste von Positionen."""
 if len(positions) < 56:
 return None
 
 chars = []
 for row, col in positions[:56]: # 56 Zeichen for Identity-Body
 if 0 <= row < 128 and 0 <= col < 128:
 chars.append(base26_to_char(matrix[row, col]))
 else:
 return None
 
 # Check ob es wie eine Identity aussieht (A-Z, 60 Zeichen)
 identity_body = ''.join(chars)
 if len(identity_body) == 56 and all(c.isupper() or c.isalpha() for c in identity_body):
 # Für jetzt: Return 56-Char Body (Checksum will später validiert)
 return identity_body
 return None

def diagonal_pattern(base_r: int, base_c: int, length: int = 14) -> List[Tuple[int, int]]:
 """Diagonal Pattern: (r+j, c+j) for j in range(length)."""
 positions = []
 for j in range(length):
 positions.append((base_r + j, base_c + j))
 return positions

def vortex_pattern(center_r: int, center_c: int, radius: int = 4) -> List[Tuple[int, int]]:
 """9-Vortex Pattern: Ring um Center."""
 positions = []
 # Vereinfachte Version: 9-Punkt Pattern
 offsets = [(-1, -1), (-1, 0), (-1, 1), (0, -1), (0, 0), (0, 1), (1, -1), (1, 0), (1, 1)]
 for dr, dc in offsets:
 positions.append((center_r + dr * radius, center_c + dc * radius))
 # Erweitere auf 56 Positionen durch Wiederholung/Erweiterung
 while len(positions) < 56:
 # Erweitere Pattern
 for dr, dc in offsets:
 if len(positions) >= 56:
 break
 positions.append((center_r + dr * (radius + len(positions) // 9), center_c + dc * (radius + len(positions) // 9)))
 return positions[:56]

def l_shape_pattern(base_r: int, base_c: int, width: int = 7, height: int = 8) -> List[Tuple[int, int]]:
 """L-Shape Pattern."""
 positions = []
 # Horizontal
 for c in range(base_c, base_c + width):
 positions.append((base_r, c))
 # Vertical
 for r in range(base_r + 1, base_r + height):
 positions.append((r, base_c + width - 1))
 # Erweitere auf 56
 while len(positions) < 56:
 positions.extend(positions[:min(56 - len(positions), len(positions))])
 return positions[:56]

def horizontal_pattern(base_r: int, base_c: int, length: int = 56) -> List[Tuple[int, int]]:
 """Horizontal Pattern: Eine Zeile."""
 return [(base_r, base_c + j) for j in range(length)]

def vertical_pattern(base_r: int, base_c: int, length: int = 56) -> List[Tuple[int, int]]:
 """Vertical Pattern: Eine Spalte."""
 return [(base_r + j, base_c) for j in range(length)]

def spiral_pattern(center_r: int, center_c: int) -> List[Tuple[int, int]]:
 """Spiral Pattern: Von außen nach innen."""
 positions = []
 # Vereinfachte Spiral
 directions = [(0, 1), (1, 0), (0, -1), (-1, 0)]
 r, c = center_r, center_c
 step = 1
 dir_idx = 0
 
 while len(positions) < 56:
 dr, dc = directions[dir_idx % 4]
 for _ in range(step):
 if len(positions) >= 56:
 break
 positions.append((r, c))
 r += dr
 c += dc
 dir_idx += 1
 if dir_idx % 2 == 0:
 step += 1
 
 return positions[:56]

def main():
 """Systematische Extraktion aller möglichen Identities."""
 
 print("=" * 80)
 print("SYSTEMATISCHE MATRIX-EXTRAKTION")
 print("=" * 80)
 print()
 
 # Load Matrix
 if not MATRIX_PATH.exists():
 print(f"❌ Matrix nicht gefunden: {MATRIX_PATH}")
 return
 
 print(f"Load Matrix: {MATRIX_PATH}")
 matrix = load_matrix(MATRIX_PATH)
 print(f"✅ Matrix geloadn: {matrix.shape}")
 print()
 
 # Bekannte Identities (zum Vergleich)
 known_identities = {
 "OBWIIPWFPOWIQCHGJHKFZNMSSYOBYNNDLUEXAELXKEROIUKIXEUXMLZBICUD",
 "OQOOMLAQLGOJFFAYGXALSTDVDGQDWKWGDQJRMNZSODHMWWWNSLSQZZAHOAZE",
 "WEZPWOMKYYQYGDZJDUEPIOTTUKCCQVBYEMYHQUTWGAMHFVJJVRCQLMVGYDGG",
 "BUICAHKIBLQWQAIONARLYROQGMRAYEAOECSZCPMTEHUIFXLKGTMAPJFDCHQI",
 "FAEEIVWNINMPFAAWUUYMCJXMSFCAGDJNFDRCEHBFPGFCCEKUWTMCBHXBNSVL",
 "PRUATXFVEFFWAEHUADBSBKOFEQTCYOXPSJHWPUSUKCHBXGMRQQEWITAELCNI",
 "RIOIMZKDJPHCFFGVYMWSFOKVBNXAYCKUXOLHLHHCPCQDULIITMFGZQUFIMKN",
 "DZGTELPKNEITGBRUPCWRNZGLTMBCRPLRPERUFBZHDFDTFYTJXOUDYLSBKRRB",
 }
 
 print("Bekannte Identities (zum Vergleich):")
 print(f" {len(known_identities)} Identities")
 print()
 
 # Systematische Extraktion
 print("=" * 80)
 print("SYSTEMATISCHE EXTRAKTION")
 print("=" * 80)
 print()
 
 all_candidates = set()
 pattern_results = defaultdict(list)
 
 # Pattern 1: Diagonal (verschiedene Start-Positionen)
 print("Pattern 1: Diagonal (verschiedene Start-Positionen)...")
 diagonal_count = 0
 for base_r in range(0, 128, 4): # Schritt 4
 for base_c in range(0, 128, 4):
 positions = diagonal_pattern(base_r, base_c, 56)
 identity = extract_identity_from_positions(matrix, positions)
 if identity:
 all_candidates.add(identity)
 pattern_results["diagonal"].append({
 "start": (base_r, base_c),
 "identity": identity,
 })
 diagonal_count += 1
 print(f" ✅ {diagonal_count} Kandidaten gefunden")
 
 # Pattern 2: Horizontal
 print("Pattern 2: Horizontal...")
 horizontal_count = 0
 for base_r in range(0, 128, 2):
 for base_c in range(0, 128 - 56):
 positions = horizontal_pattern(base_r, base_c, 56)
 identity = extract_identity_from_positions(matrix, positions)
 if identity:
 all_candidates.add(identity)
 pattern_results["horizontal"].append({
 "start": (base_r, base_c),
 "identity": identity,
 })
 horizontal_count += 1
 print(f" ✅ {horizontal_count} Kandidaten gefunden")
 
 # Pattern 3: Vertical
 print("Pattern 3: Vertical...")
 vertical_count = 0
 for base_r in range(0, 128 - 56):
 for base_c in range(0, 128, 2):
 positions = vertical_pattern(base_r, base_c, 56)
 identity = extract_identity_from_positions(matrix, positions)
 if identity:
 all_candidates.add(identity)
 pattern_results["vertical"].append({
 "start": (base_r, base_c),
 "identity": identity,
 })
 vertical_count += 1
 print(f" ✅ {vertical_count} Kandidaten gefunden")
 
 # Pattern 4: L-Shape
 print("Pattern 4: L-Shape...")
 lshape_count = 0
 for base_r in range(0, 128 - 8):
 for base_c in range(0, 128 - 7):
 positions = l_shape_pattern(base_r, base_c)
 identity = extract_identity_from_positions(matrix, positions)
 if identity:
 all_candidates.add(identity)
 pattern_results["lshape"].append({
 "start": (base_r, base_c),
 "identity": identity,
 })
 lshape_count += 1
 print(f" ✅ {lshape_count} Kandidaten gefunden")
 
 # Pattern 5: Spiral
 print("Pattern 5: Spiral...")
 spiral_count = 0
 for center_r in range(10, 118, 10):
 for center_c in range(10, 118, 10):
 positions = spiral_pattern(center_r, center_c)
 identity = extract_identity_from_positions(matrix, positions)
 if identity:
 all_candidates.add(identity)
 pattern_results["spiral"].append({
 "center": (center_r, center_c),
 "identity": identity,
 })
 spiral_count += 1
 print(f" ✅ {spiral_count} Kandidaten gefunden")
 
 print()
 print("=" * 80)
 print("ZUSAMMENFASSUNG")
 print("=" * 80)
 print()
 
 print(f"Total Kandidaten gefunden: {len(all_candidates)}")
 print()
 print("Nach Pattern:")
 for pattern, results in pattern_results.items():
 print(f" {pattern}: {len(results)} Kandidaten")
 
 # Check bekannte Identities
 print()
 print("Bekannte Identities gefunden:")
 found_known = 0
 for known_id in known_identities:
 # Check ob Body (erste 56 Zeichen) in Kandidaten
 known_body = known_id[:56]
 if known_body in all_candidates:
 found_known += 1
 print(f" ✅ {known_id[:40]}...")
 
 print(f" {found_known}/{len(known_identities)} bekannte Identities gefunden")
 print()
 
 # Speichere Ergebnisse
 OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
 
 # Speichere ALLE Kandidaten (nicht nur Stichprobe)
 all_candidates_list = sorted(list(all_candidates))
 
 results = {
 "summary": {
 "total_candidates": len(all_candidates),
 "known_found": found_known,
 "patterns_tested": len(pattern_results),
 },
 "pattern_results": {
 pattern: {
 "count": len(results),
 "candidates": [r["identity"] for r in results] # ALLE Kandidaten
 }
 for pattern, results in pattern_results.items()
 },
 "all_candidates": all_candidates_list, # ALLE Kandidaten
 }
 
 # Speichere in JSON (wenn < 50MB, sonst Batch-Speicherung)
 output_file = OUTPUT_DIR / "systematic_matrix_extraction.json"
 
 # Check Größe (ca. 60 Bytes pro Kandidat)
 estimated_size = len(all_candidates_list) * 60
 if estimated_size > 50 * 1024 * 1024: # > 50MB
 print(f"⚠️ Datei will groß (~{estimated_size / 1024 / 1024:.1f} MB)")
 print(" Verwende Batch-Speicherung...")
 # Speichere in mehreren Dateien
 batch_size = 10000
 for i in range(0, len(all_candidates_list), batch_size):
 batch = all_candidates_list[i:i+batch_size]
 batch_file = OUTPUT_DIR / f"systematic_matrix_extraction_batch_{i//batch_size}.json"
 with batch_file.open("w") as f:
 json.dump(batch, f, indent=2)
 print(f" Batch {i//batch_size + 1} gespeichert: {len(batch)} Kandidaten")
 
 # Speichere Summary separat
 summary_data = {
 "summary": results["summary"],
 "pattern_results": results["pattern_results"],
 "total_batches": (len(all_candidates_list) + batch_size - 1) // batch_size,
 "batch_size": batch_size,
 }
 with output_file.open("w") as f:
 json.dump(summary_data, f, indent=2)
 else:
 with output_file.open("w") as f:
 json.dump(results, f, indent=2)
 
 print(f"💾 Ergebnisse gespeichert in: {OUTPUT_DIR / 'systematic_matrix_extraction.json'}")
 print()
 print("Nächster Schritt: On-Chain Validierung der Kandidaten!")

if __name__ == "__main__":
 main()

