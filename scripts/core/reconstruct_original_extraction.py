#!/usr/bin/env python3
"""
Rekonstruiere die Original-Extraktionsmethoden for die 8 bekannten Identities.

Analysiert die bekannten Identities und findet die exakten Extraktionsmethoden.
"""

import json
import numpy as np
from pathlib import Path
from typing import List, Tuple, Optional, Dict

OUTPUT_DIR = Path("outputs/derived")
MATRIX_PATH = Path("data/anna-matrix/Anna_Matrix.xlsx")

# Die 8 bekannten Layer-1 Identities
KNOWN_IDENTITIES = [
 "OBWIIPWFPOWIQCHGJHKFZNMSSYOBYNNDLUEXAELXKEROIUKIXEUXMLZBICUD", # Diagonal #1
 "OQOOMLAQLGOJFFAYGXALSTDVDGQDWKWGDQJRMNZSODHMWWWNSLSQZZAHOAZE", # Diagonal #2
 "WEZPWOMKYYQYGDZJDUEPIOTTUKCCQVBYEMYHQUTWGAMHFVJJVRCQLMVGYDGG", # Diagonal #3
 "BUICAHKIBLQWQAIONARLYROQGMRAYEAOECSZCPMTEHUIFXLKGTMAPJFDCHQI", # Diagonal #4
 "FAEEIVWNINMPFAAWUUYMCJXMSFCAGDJNFDRCEHBFPGFCCEKUWTMCBHXBNSVL", # Vortex #1
 "PRUATXFVEFFWAEHUADBSBKOFEQTCYOXPSJHWPUSUKCHBXGMRQQEWITAELCNI", # Vortex #2
 "RIOIMZKDJPHCFFGVYMWSFOKVBNXAYCKUXOLHLHHCPCQDULIITMFGZQUFIMKN", # Vortex #3
 "DZGTELPKNEITGBRUPCWRNZGLTMBCRPLRPERUFBZHDFDTFYTJXOUDYLSBKRRB", # Vortex #4
]

def load_matrix(path: Path) -> np.ndarray:
 """Load Matrix aus Excel mit openpyxl."""
 try:
 from openpyxl import load_workbook
 wb = load_workbook(path, data_only=True)
 ws = wb.active
 matrix = np.zeros((128, 128), dtype=float)
 for i in range(128):
 for j in range(128):
 val = ws.cell(i+1, j+1).value
 try:
 matrix[i, j] = float(val) if val is not None else 0.0
 except (ValueError, TypeError):
 matrix[i, j] = 0.0
 return matrix
 except ImportError:
 print("ERROR: openpyxl nicht installiert. Installiere mit: pip3 install openpyxl")
 raise

def base26_to_char(value: float) -> str:
 """Konvertiere Matrix-Wert zu Base-26 Charakter."""
 return chr(ord('A') + (int(value) % 26))

def extract_from_positions(matrix: np.ndarray, positions: List[Tuple[int, int]]) -> str:
 """Extrahiere String aus Positionen."""
 chars = []
 for row, col in positions:
 if 0 <= row < 128 and 0 <= col < 128:
 chars.append(base26_to_char(matrix[row, col]))
 else:
 return None
 return ''.join(chars)

def find_diagonal_pattern(matrix: np.ndarray, target_body: str) -> Optional[Dict]:
 """Finde Diagonal-Pattern, das zu target_body passt."""
 # Basierend auf COMPLETE_PROOF.md: diagonal (r+j, c+j) in 32×32 windows
 # for group_start in [0, 32, 64, 96]: # 4 identities
 # for group in range(4):
 # row_offset = group_start + (group // 2) * 16
 # col_offset = (group % 2) * 16
 # for j in range(14):
 # row = row_offset + j
 # col = col_offset + j
 
 for group_start in [0, 32, 64, 96]:
 for group in range(4):
 row_offset = group_start + (group // 2) * 16
 col_offset = (group % 2) * 16
 
 # Extrahiere 4 Blöcke à 14 Zeichen = 56 Zeichen
 positions = []
 for block in range(4):
 base_r = row_offset + (block // 2) * 16
 base_c = col_offset + (block % 2) * 16
 for j in range(14):
 positions.append((base_r + j, base_c + j))
 
 if len(positions) >= 56:
 extracted = extract_from_positions(matrix, positions[:56])
 if extracted and extracted == target_body:
 return {
 "method": "diagonal",
 "group_start": group_start,
 "group": group,
 "row_offset": row_offset,
 "col_offset": col_offset,
 "positions": positions[:56],
 }
 return None

def find_vortex_pattern(matrix: np.ndarray, target_body: str) -> Optional[Dict]:
 """Finde Vortex-Pattern (9-Vortex Rings)."""
 # Basierend auf STRUCTURE_DISCOVERY_SUMMARY: 9-Vortex rings
 # Teste verschiedene Center-Positionen und Radien
 
 for center_r in range(10, 118, 4):
 for center_c in range(10, 118, 4):
 for radius in [1, 2, 3, 4, 5]:
 # 9-Vortex Pattern: 3x3 Grid um Center
 positions = []
 offsets = [(-1, -1), (-1, 0), (-1, 1), (0, -1), (0, 0), (0, 1), (1, -1), (1, 0), (1, 1)]
 
 # Erweitere auf 56 Positionen durch mehrere Ringe
 for ring in range(4): # 4 Ringe à 9 Punkte = 36, dann erweitern
 for dr, dc in offsets:
 if len(positions) >= 56:
 break
 r = center_r + dr * (radius + ring)
 c = center_c + dc * (radius + ring)
 positions.append((r, c))
 
 # Fülle auf 56 auf
 while len(positions) < 56:
 positions.extend(positions[:min(56 - len(positions), len(positions))])
 
 if len(positions) >= 56:
 extracted = extract_from_positions(matrix, positions[:56])
 if extracted and extracted == target_body:
 return {
 "method": "vortex",
 "center": (center_r, center_c),
 "radius": radius,
 "positions": positions[:56],
 }
 return None

def main():
 """Rekonstruiere Original-Extraktionsmethoden."""
 
 print("=" * 80)
 print("REKONSTRUKTION DER ORIGINAL-EXTRAKTIONSMETHODEN")
 print("=" * 80)
 print()
 
 # Load Matrix
 print(f"Load Matrix: {MATRIX_PATH}")
 matrix = load_matrix(MATRIX_PATH)
 print(f"✅ Matrix geloadn: {matrix.shape}")
 print()
 
 # Extrahiere Bodies (erste 56 Zeichen)
 known_bodies = {id[:56]: id for id in KNOWN_IDENTITIES}
 
 print(f"Suche Extraktionsmethoden for {len(KNOWN_IDENTITIES)} bekannte Identities...")
 print()
 
 results = {}
 
 # Teste Diagonal-Patterns (erste 4)
 print("Teste Diagonal-Patterns...")
 for i, identity in enumerate(KNOWN_IDENTITIES[:4], 1):
 body = identity[:56]
 print(f" Diagonal #{i}: {body[:30]}...")
 
 pattern = find_diagonal_pattern(matrix, body)
 if pattern:
 results[identity] = pattern
 print(f" ✅ Gefunden: group_start={pattern['group_start']}, group={pattern['group']}")
 else:
 print(f" ❌ Nicht gefunden")
 print()
 
 # Teste Vortex-Patterns (letzte 4)
 print("Teste Vortex-Patterns...")
 for i, identity in enumerate(KNOWN_IDENTITIES[4:], 1):
 body = identity[:56]
 print(f" Vortex #{i}: {body[:30]}...")
 
 pattern = find_vortex_pattern(matrix, body)
 if pattern:
 results[identity] = pattern
 print(f" ✅ Gefunden: center={pattern['center']}, radius={pattern['radius']}")
 else:
 print(f" ❌ Nicht gefunden")
 print()
 
 # Zusammenfassung
 print("=" * 80)
 print("ZUSAMMENFASSUNG")
 print("=" * 80)
 print()
 print(f"Gefunden: {len(results)}/{len(KNOWN_IDENTITIES)} Identities")
 print()
 
 # Speichere Ergebnisse
 OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
 output_data = {
 "summary": {
 "total_identities": len(KNOWN_IDENTITIES),
 "found": len(results),
 "success_rate": len(results) / len(KNOWN_IDENTITIES) * 100,
 },
 "results": results,
 }
 
 output_file = OUTPUT_DIR / "original_extraction_methods.json"
 with output_file.open("w") as f:
 json.dump(output_data, f, indent=2)
 
 print(f"💾 Ergebnisse gespeichert in: {output_file}")
 
 if len(results) == len(KNOWN_IDENTITIES):
 print()
 print("✅ Alle bekannten Identities gefunden!")
 print(" Nächster Schritt: Verwende diese Methoden, um weitere Identities zu finden.")
 else:
 print()
 print("⚠️ Nicht alle Identities gefunden.")
 print(" Mögliche Gründe:")
 print(" - Pattern-Parameter müssen angepasst werden")
 print(" - Andere Extraktionsmethoden verwendet")

if __name__ == "__main__":
 main()

