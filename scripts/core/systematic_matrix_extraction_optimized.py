#!/usr/bin/env python3
"""
Optimierte Systematische Matrix-Extraktion mit Checkpoint-System.

Features:
- Batch-Processing for Performance
- Checkpoint-System for Unterbrechungen
- Progress-Tracking
- Effiziente Datenstrukturen
- Vollständige Speicherung aller Kandidaten
"""

import json
import numpy as np
from pathlib import Path
from typing import List, Set, Tuple, Optional, Dict
from collections import defaultdict
import time

OUTPUT_DIR = Path("outputs/derived")
MATRIX_PATH = Path("data/anna-matrix/Anna_Matrix.xlsx")
CHECKPOINT_FILE = OUTPUT_DIR / "matrix_extraction_checkpoint.json"
OUTPUT_FILE = OUTPUT_DIR / "systematic_matrix_extraction_complete.json"
BATCH_SIZE = 10000 # Speichere in Batches von 10.000

def load_matrix(path: Path) -> np.ndarray:
 """Load Matrix aus Excel mit openpyxl."""
 try:
 from openpyxl import load_workbook
 wb = load_workbook(path, data_only=True)
 ws = wb.active
 matrix = np.zeros((128, 128), dtype=float)
 for i in range(128):
 for j in range(128):
 val = ws.cell(i+1, j+1).value
 try:
 matrix[i, j] = float(val) if val is not None else 0.0
 except (ValueError, TypeError):
 matrix[i, j] = 0.0
 return matrix
 except ImportError:
 print("ERROR: openpyxl nicht installiert. Installiere mit: pip3 install openpyxl")
 raise

def base26_to_char(value: float) -> str:
 """Konvertiere Matrix-Wert zu Base-26 Charakter."""
 return chr(ord('A') + (int(value) % 26))

def extract_identity_from_positions(matrix: np.ndarray, positions: List[Tuple[int, int]]) -> Optional[str]:
 """Extrahiere Identity aus Liste von Positionen."""
 if len(positions) < 56:
 return None
 
 chars = []
 for row, col in positions[:56]:
 if 0 <= row < 128 and 0 <= col < 128:
 chars.append(base26_to_char(matrix[row, col]))
 else:
 return None
 
 identity_body = ''.join(chars)
 if len(identity_body) == 56 and all(c.isupper() or c.isalpha() for c in identity_body):
 return identity_body
 return None

def diagonal_pattern(base_r: int, base_c: int, length: int = 14) -> List[Tuple[int, int]]:
 """Diagonal Pattern: (r+j, c+j) for j in range(length)."""
 return [(base_r + j, base_c + j) for j in range(length)]

def horizontal_pattern(base_r: int, base_c: int, length: int = 56) -> List[Tuple[int, int]]:
 """Horizontal Pattern: Eine Zeile."""
 return [(base_r, base_c + j) for j in range(length)]

def vertical_pattern(base_r: int, base_c: int, length: int = 56) -> List[Tuple[int, int]]:
 """Vertical Pattern: Eine Spalte."""
 return [(base_r + j, base_c) for j in range(length)]

def l_shape_pattern(base_r: int, base_c: int, width: int = 7, height: int = 8) -> List[Tuple[int, int]]:
 """L-Shape Pattern."""
 positions = []
 # Horizontal
 for c in range(base_c, base_c + width):
 positions.append((base_r, c))
 # Vertical
 for r in range(base_r + 1, base_r + height):
 positions.append((r, base_c + width - 1))
 # Erweitere auf 56
 while len(positions) < 56:
 positions.extend(positions[:min(56 - len(positions), len(positions))])
 return positions[:56]

def spiral_pattern(center_r: int, center_c: int) -> List[Tuple[int, int]]:
 """Spiral Pattern: Von außen nach innen."""
 positions = []
 directions = [(0, 1), (1, 0), (0, -1), (-1, 0)]
 r, c = center_r, center_c
 step = 1
 dir_idx = 0
 
 while len(positions) < 56:
 dr, dc = directions[dir_idx % 4]
 for _ in range(step):
 if len(positions) >= 56:
 break
 positions.append((r, c))
 r += dr
 c += dc
 dir_idx += 1
 if dir_idx % 2 == 0:
 step += 1
 
 return positions[:56]

def load_checkpoint() -> Dict:
 """Load Checkpoint falls vorhanden."""
 if CHECKPOINT_FILE.exists():
 try:
 with CHECKPOINT_FILE.open() as f:
 return json.load(f)
 except Exception as e:
 print(f"⚠️ Checkpoint konnte nicht geloadn werden: {e}")
 return {
 "all_candidates": set(),
 "pattern_results": defaultdict(list),
 "patterns_completed": set(),
 "last_pattern": None,
 "last_position": None,
 }

def save_checkpoint(checkpoint: Dict):
 """Speichere Checkpoint."""
 OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
 # Konvertiere Set zu List for JSON
 checkpoint_save = {
 "all_candidates": sorted(list(checkpoint["all_candidates"])),
 "pattern_results": {
 pattern: results
 for pattern, results in checkpoint["pattern_results"].items()
 },
 "patterns_completed": list(checkpoint["patterns_completed"]),
 "last_pattern": checkpoint.get("last_pattern"),
 "last_position": checkpoint.get("last_position"),
 }
 with CHECKPOINT_FILE.open("w") as f:
 json.dump(checkpoint_save, f, indent=2)

def extract_pattern_optimized(
 matrix: np.ndarray,
 pattern_func,
 pattern_name: str,
 start_params: List[Tuple],
 checkpoint: Dict
) -> Tuple[Set[str], List[Dict]]:
 """Extrahiere Kandidaten for ein Pattern (optimiert)."""
 
 # Check ob Pattern bereits komplett
 if pattern_name in checkpoint["patterns_completed"]:
 print(f" ⏭️ Pattern '{pattern_name}' bereits komplett (Checkpoint)")
 return set(), []
 
 candidates = set()
 results = []
 
 # Check ob wir bei diesem Pattern weitermachen müssen
 resume_from = None
 if checkpoint.get("last_pattern") == pattern_name and checkpoint.get("last_position"):
 resume_from = checkpoint["last_position"]
 print(f" 🔄 Setze fort ab Position: {resume_from}")
 
 start_idx = 0
 if resume_from is not None:
 # Finde Start-Index
 for i, params in enumerate(start_params):
 if params == resume_from:
 start_idx = i
 break
 
 total = len(start_params)
 for i, params in enumerate(start_params[start_idx:], start_idx):
 if i % 1000 == 0 and i > 0:
 print(f" Progress: {i}/{total} ({i/total*100:.1f}%)")
 # Speichere Checkpoint alle 1000 Iterationen
 checkpoint["last_pattern"] = pattern_name
 checkpoint["last_position"] = params
 save_checkpoint(checkpoint)
 
 positions = pattern_func(*params) if isinstance(params, tuple) else pattern_func(params)
 identity = extract_identity_from_positions(matrix, positions)
 
 if identity:
 candidates.add(identity)
 results.append({
 "params": params,
 "identity": identity,
 })
 
 # Markiere Pattern als komplett
 checkpoint["patterns_completed"].add(pattern_name)
 checkpoint["last_pattern"] = None
 checkpoint["last_position"] = None
 
 return candidates, results

def main():
 """Optimierte systematische Extraktion mit Checkpoint-System."""
 
 print("=" * 80)
 print("OPTIMIERTE SYSTEMATISCHE MATRIX-EXTRAKTION")
 print("=" * 80)
 print()
 
 # Load Checkpoint
 checkpoint = load_checkpoint()
 if checkpoint.get("all_candidates"):
 print(f"✅ Checkpoint geloadn: {len(checkpoint['all_candidates'])} Kandidaten bereits gefunden")
 print(f" Patterns komplett: {len(checkpoint['patterns_completed'])}")
 print()
 
 # Load Matrix
 print(f"Load Matrix: {MATRIX_PATH}")
 matrix = load_matrix(MATRIX_PATH)
 print(f"✅ Matrix geloadn: {matrix.shape}")
 print()
 
 # Bekannte Identities (zum Vergleich)
 known_identities = [
 "OBWIIPWFPOWIQCHGJHKFZNMSSYOBYNNDLUEXAELXKEROIUKIXEUXMLZBICUD",
 "OQOOMLAQLGOJFFAYGXALSTDVDGQDWKWGDQJRMNZSODHMWWWNSLSQZZAHOAZE",
 "WEZPWOMKYYQYGDZJDUEPIOTTUKCCQVBYEMYHQUTWGAMHFVJJVRCQLMVGYDGG",
 "BUICAHKIBLQWQAIONARLYROQGMRAYEAOECSZCPMTEHUIFXLKGTMAPJFDCHQI",
 "FAEEIVWNINMPFAAWUUYMCJXMSFCAGDJNFDRCEHBFPGFCCEKUWTMCBHXBNSVL",
 "PRUATXFVEFFWAEHUADBSBKOFEQTCYOXPSJHWPUSUKCHBXGMRQQEWITAELCNI",
 "RIOIMZKDJPHCFFGVYMWSFOKVBNXAYCKUXOLHLHHCPCQDULIITMFGZQUFIMKN",
 "DZGTELPKNEITGBRUPCWRNZGLTMBCRPLRPERUFBZHDFDTFYTJXOUDYLSBKRRB",
 ]
 known_bodies = {id[:56] for id in known_identities}
 
 # Initialisiere Checkpoint
 if not checkpoint.get("all_candidates"):
 checkpoint["all_candidates"] = set()
 checkpoint["pattern_results"] = defaultdict(list)
 checkpoint["patterns_completed"] = set()
 
 all_candidates = checkpoint["all_candidates"]
 pattern_results = checkpoint["pattern_results"]
 
 # Pattern 1: Diagonal (optimiert - weniger Iterationen)
 print("Pattern 1: Diagonal (optimiert)...")
 start_time = time.time()
 diagonal_params = [(r, c) for r in range(0, 128, 4) for c in range(0, 128, 4)]
 diag_candidates, diag_results = extract_pattern_optimized(
 matrix, diagonal_pattern, "diagonal", diagonal_params, checkpoint
 )
 all_candidates.update(diag_candidates)
 pattern_results["diagonal"].extend(diag_results)
 print(f" ✅ {len(diag_candidates)} Kandidaten gefunden ({time.time() - start_time:.1f}s)")
 save_checkpoint(checkpoint)
 print()
 
 # Pattern 2: Horizontal (optimiert)
 print("Pattern 2: Horizontal (optimiert)...")
 start_time = time.time()
 horizontal_params = [(r, c) for r in range(0, 128, 2) for c in range(0, 128 - 56)]
 horiz_candidates, horiz_results = extract_pattern_optimized(
 matrix, horizontal_pattern, "horizontal", horizontal_params, checkpoint
 )
 all_candidates.update(horiz_candidates)
 pattern_results["horizontal"].extend(horiz_results)
 print(f" ✅ {len(horiz_candidates)} Kandidaten gefunden ({time.time() - start_time:.1f}s)")
 save_checkpoint(checkpoint)
 print()
 
 # Pattern 3: Vertical (optimiert)
 print("Pattern 3: Vertical (optimiert)...")
 start_time = time.time()
 vertical_params = [(r, c) for r in range(0, 128 - 56) for c in range(0, 128, 2)]
 vert_candidates, vert_results = extract_pattern_optimized(
 matrix, vertical_pattern, "vertical", vertical_params, checkpoint
 )
 all_candidates.update(vert_candidates)
 pattern_results["vertical"].extend(vert_results)
 print(f" ✅ {len(vert_candidates)} Kandidaten gefunden ({time.time() - start_time:.1f}s)")
 save_checkpoint(checkpoint)
 print()
 
 # Pattern 4: L-Shape (optimiert)
 print("Pattern 4: L-Shape (optimiert)...")
 start_time = time.time()
 lshape_params = [(r, c) for r in range(0, 128 - 8) for c in range(0, 128 - 7)]
 lshape_candidates, lshape_results = extract_pattern_optimized(
 matrix, l_shape_pattern, "lshape", lshape_params, checkpoint
 )
 all_candidates.update(lshape_candidates)
 pattern_results["lshape"].extend(lshape_results)
 print(f" ✅ {len(lshape_candidates)} Kandidaten gefunden ({time.time() - start_time:.1f}s)")
 save_checkpoint(checkpoint)
 print()
 
 # Pattern 5: Spiral (optimiert)
 print("Pattern 5: Spiral (optimiert)...")
 start_time = time.time()
 spiral_params = [(r, c) for r in range(10, 118, 10) for c in range(10, 118, 10)]
 spiral_candidates, spiral_results = extract_pattern_optimized(
 matrix, spiral_pattern, "spiral", spiral_params, checkpoint
 )
 all_candidates.update(spiral_candidates)
 pattern_results["spiral"].extend(spiral_results)
 print(f" ✅ {len(spiral_candidates)} Kandidaten gefunden ({time.time() - start_time:.1f}s)")
 save_checkpoint(checkpoint)
 print()
 
 # Zusammenfassung
 print("=" * 80)
 print("ZUSAMMENFASSUNG")
 print("=" * 80)
 print()
 
 print(f"Total Kandidaten gefunden: {len(all_candidates):,}")
 print()
 print("Nach Pattern:")
 for pattern, results in pattern_results.items():
 print(f" {pattern}: {len(results):,} Kandidaten")
 
 # Check bekannte Identities
 print()
 print("Bekannte Identities gefunden:")
 found_known = 0
 for known_body in known_bodies:
 if known_body in all_candidates:
 found_known += 1
 
 print(f" {found_known}/{len(known_identities)} bekannte Identities gefunden")
 print()
 
 # Speichere finale Ergebnisse
 print("Speichere finale Ergebnisse...")
 all_candidates_list = sorted(list(all_candidates))
 
 # Batch-Speicherung falls nötig
 if len(all_candidates_list) > BATCH_SIZE:
 print(f" Datei will groß ({len(all_candidates_list):,} Kandidaten)")
 print(f" Verwende Batch-Speicherung (Batch-Size: {BATCH_SIZE:,})...")
 
 # Speichere in mehreren Dateien
 num_batches = (len(all_candidates_list) + BATCH_SIZE - 1) // BATCH_SIZE
 for i in range(num_batches):
 batch = all_candidates_list[i * BATCH_SIZE:(i + 1) * BATCH_SIZE]
 batch_file = OUTPUT_DIR / f"matrix_candidates_batch_{i}.json"
 with batch_file.open("w") as f:
 json.dump(batch, f, indent=2)
 print(f" Batch {i+1}/{num_batches} gespeichert: {len(batch):,} Kandidaten")
 
 # Speichere Summary
 summary_data = {
 "summary": {
 "total_candidates": len(all_candidates),
 "known_found": found_known,
 "patterns_tested": len(pattern_results),
 },
 "pattern_results": {
 pattern: {
 "count": len(results),
 "sample": [r["identity"] for r in results[:10]] # Stichprobe for Übersicht
 }
 for pattern, results in pattern_results.items()
 },
 "total_batches": num_batches,
 "batch_size": BATCH_SIZE,
 }
 with OUTPUT_FILE.open("w") as f:
 json.dump(summary_data, f, indent=2)
 else:
 # Normale Speicherung
 results = {
 "summary": {
 "total_candidates": len(all_candidates),
 "known_found": found_known,
 "patterns_tested": len(pattern_results),
 },
 "pattern_results": {
 pattern: {
 "count": len(results),
 "candidates": [r["identity"] for r in results]
 }
 for pattern, results in pattern_results.items()
 },
 "all_candidates": all_candidates_list,
 }
 with OUTPUT_FILE.open("w") as f:
 json.dump(results, f, indent=2)
 
 print(f"💾 Ergebnisse gespeichert in: {OUTPUT_FILE}")
 
 # Lösche Checkpoint (Extraktion komplett)
 if CHECKPOINT_FILE.exists():
 CHECKPOINT_FILE.unlink()
 print("✅ Checkpoint gelöscht (Extraktion komplett)")
 
 print()
 print("Nächster Schritt: Checksum-Berechnung for alle Kandidaten!")

if __name__ == "__main__":
 main()

